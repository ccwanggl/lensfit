"""Excel report generator for LensFit match results."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_report(
    requirements: dict[str, Any],
    results: list[dict[str, Any]],
    top_k: int = 20,
    diagnostics: list[dict[str, Any]] | None = None,
    what_if_results: list[dict[str, Any]] | None = None,
) -> bytes:
    """Generate an Excel report of match results.

    Args:
        requirements: User input parameters.
        results: Top match results.
        top_k: Max results to include.
        diagnostics: Filter stage diagnostics (optional).
        what_if_results: What-if sensitivity results (optional).

    Returns:
        Excel bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "选型报告"

    # Styles
    title_font = Font(size=16, bold=True, color="1A56DB")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    def _sanitize(val):
        """防 Excel 公式注入：以 = + - @ 开头的字符串前加单引号."""
        if isinstance(val, str) and val and val[0] in "=+-@":
            return "'" + val
        return val

    # Title
    ws["A1"] = "LensFit 选型报告"
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")

    # Requirements
    ws["A3"] = "选型参数"
    ws["A3"].font = Font(size=12, bold=True)
    req_items = [
        ("传感器尺寸", requirements.get("sensor_size", "-")),
        ("像元尺寸 (μm)", requirements.get("pixel_size_um", "-")),
        ("目标宽度 (mm)", requirements.get("target_width_mm", "-")),
        ("目标高度 (mm)", requirements.get("target_height_mm", "-")),
        ("工作距离 (mm)", requirements.get("working_distance_mm", "-")),
        ("镜头类型", requirements.get("lens_type", "-")),
        ("接口类型", requirements.get("interface", "-")),
    ]
    for i, (k, v) in enumerate(req_items, start=4):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = _sanitize(v)
        ws[f"A{i}"].font = Font(bold=True)

    # Results header
    result_start = 13
    headers = ["排名", "镜头型号", "探测器型号", "评分", "覆盖度", "渐晕", "估算焦距 (mm)", "放大倍率", "匹配理由"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=result_start, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Results data
    for idx, r in enumerate(results[:top_k], start=1):
        row = result_start + idx
        derived = r.get("derived", {})
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=_sanitize(r.get("lens_model", f"#{r.get('lens_id', '?')}")))
        ws.cell(row=row, column=3, value=_sanitize(r.get("detector_model", f"#{r.get('detector_id', '?')}")))
        ws.cell(row=row, column=4, value=round(r.get("score", 0), 3))
        ws.cell(row=row, column=5, value=f"{(r.get('coverage_ratio', 0) * 100):.0f}%")
        ws.cell(row=row, column=6, value="是" if r.get("vignetting") else "否")
        ws.cell(row=row, column=7, value=derived.get("focal_length", "-"))
        ws.cell(row=row, column=8, value=derived.get("magnification", "-"))
        ws.cell(row=row, column=9, value=_sanitize(r.get("reason", "-")))

        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if idx % 2 == 0:
                cell.fill = alt_fill

    # Auto column width
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["I"].width = 40

    # ─── Diagnostics sheet ───
    if diagnostics:
        ws_diag = wb.create_sheet(title="诊断")
        ws_diag["A1"] = "匹配诊断"
        ws_diag["A1"].font = Font(size=14, bold=True, color="1A56DB")
        ws_diag.merge_cells("A1:E1")

        diag_headers = ["阶段", "过滤前", "过滤后", "拒绝原因", "建议"]
        for col, h in enumerate(diag_headers, start=1):
            cell = ws_diag.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for idx, d in enumerate(diagnostics, start=1):
            row = 3 + idx
            reasons = d.get("rejected_reasons", {})
            reason_str = "; ".join(f"{k}: {v}" for k, v in reasons.items()) if reasons else "-"
            ws_diag.cell(row=row, column=1, value=d.get("stage", "-"))
            ws_diag.cell(row=row, column=2, value=d.get("before_count", 0))
            ws_diag.cell(row=row, column=3, value=d.get("after_count", 0))
            ws_diag.cell(row=row, column=4, value=_sanitize(reason_str))
            ws_diag.cell(row=row, column=5, value=_sanitize(d.get("suggestion", "-")))
            for col in range(1, 6):
                cell = ws_diag.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if idx % 2 == 0:
                    cell.fill = alt_fill

        for col in range(1, 6):
            ws_diag.column_dimensions[get_column_letter(col)].width = 24
        ws_diag.column_dimensions["D"].width = 40
        ws_diag.column_dimensions["E"].width = 40

    # ─── Derivation chain sheet ───
    chain_exist = any(r.get("derivation_chain") for r in results[:top_k])
    if chain_exist:
        ws_chain = wb.create_sheet(title="推导链")
        ws_chain["A1"] = "推导链"
        ws_chain["A1"].font = Font(size=14, bold=True, color="1A56DB")
        ws_chain.merge_cells("A1:F1")

        chain_headers = ["方案", "步骤", "公式", "输入", "输出", "物理原理"]
        for col, h in enumerate(chain_headers, start=1):
            cell = ws_chain.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        row = 4
        for r in results[:top_k]:
            chain = r.get("derivation_chain", [])
            if not chain:
                continue
            lens_name = r.get("lens_model", f"#{r.get('lens_id', '?')}")
            for step in chain:
                ws_chain.cell(row=row, column=1, value=_sanitize(lens_name))
                ws_chain.cell(row=row, column=2, value=step.get("step", "-"))
                ws_chain.cell(row=row, column=3, value=_sanitize(step.get("formula", "-")))
                ws_chain.cell(row=row, column=4, value=_sanitize(str(step.get("inputs", "-"))[:200]))
                ws_chain.cell(row=row, column=5, value=_sanitize(str(step.get("output", "-"))[:100]))
                ws_chain.cell(row=row, column=6, value=_sanitize(step.get("principle", "-")))
                for col in range(1, 7):
                    cell = ws_chain.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                row += 1

        for col in range(1, 7):
            ws_chain.column_dimensions[get_column_letter(col)].width = 20
        ws_chain.column_dimensions["A"].width = 28
        ws_chain.column_dimensions["D"].width = 35
        ws_chain.column_dimensions["F"].width = 30

    # ─── What-if sensitivity sheet ───
    if what_if_results and results:
        ws_what = wb.create_sheet(title="敏感性")
        ws_what["A1"] = "参数敏感性分析"
        ws_what["A1"].font = Font(size=14, bold=True, color="1A56DB")
        ws_what.merge_cells("A1:E1")

        ws_what["A3"] = f"结果数量变化：{len(what_if_results) - len(results):+d} 个"
        ws_what["A3"].font = Font(size=11, bold=True)

        diff_headers = ["镜头", "基准评分", "调整后评分", "变化", "变化率"]
        for col, h in enumerate(diff_headers, start=1):
            cell = ws_what.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        baseline_map = {f"{b.get('lens_id')}-{b.get('detector_id')}": b for b in results[:top_k]}
        for idx, r in enumerate(what_if_results[:top_k], start=1):
            row = 5 + idx
            key = f"{r.get('lens_id')}-{r.get('detector_id')}"
            baseline = baseline_map.get(key)
            score = r.get("score", 0)
            base_score = baseline.get("score", 0) if baseline else 0
            diff = score - base_score
            diff_pct = (diff / base_score * 100) if base_score else 0

            ws_what.cell(row=row, column=1, value=_sanitize(r.get("lens_model", key)))
            ws_what.cell(row=row, column=2, value=round(base_score, 3) if baseline else "-")
            ws_what.cell(row=row, column=3, value=round(score, 3))
            ws_what.cell(row=row, column=4, value=round(diff, 3))
            ws_what.cell(row=row, column=5, value=f"{diff_pct:+.1f}%" if baseline else "-")

            for col in range(1, 6):
                cell = ws_what.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if idx % 2 == 0:
                    cell.fill = alt_fill

        for col in range(1, 6):
            ws_what.column_dimensions[get_column_letter(col)].width = 20
        ws_what.column_dimensions["A"].width = 32

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
