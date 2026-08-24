"""Bulk import pipeline for user-uploaded lens/detector catalogs."""

from __future__ import annotations

import csv
import io
from typing import Any

from fastapi import HTTPException
from sqlalchemy.orm import Session

from optibench.db.models import DetectorCatalog, LensCatalog, Manufacturer


def _manufacturer_id_for_name(session: Session, name: str | None) -> int | None:
    if not name:
        return None
    name = name.strip()
    if not name:
        return None
    existing = (
        session.query(Manufacturer)
        .filter(
            (Manufacturer.name.ilike(name))
            | (Manufacturer.name_en.ilike(name))
            | (Manufacturer.name_cn.ilike(name))
        )
        .first()
    )
    if existing:
        return existing.id
    item = Manufacturer(name=name, is_verified=False, data_source="user")
    session.add(item)
    session.flush()
    return item.id


def _float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _str(val: Any) -> str | None:
    if val is None:
        return None
    val = str(val).strip()
    return val if val else None


def _validated_manufacturer_id(session: Session, val: Any) -> int | None:
    m_id = _int(val)
    if m_id is None:
        return None
    if session.get(Manufacturer, m_id) is None:
        raise HTTPException(
            status_code=400,
            detail=f"manufacturer_id {m_id} does not exist",
        )
    return m_id


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _rows_from_excel(content: bytes, max_rows: int = 10000) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(
            status_code=400, detail="Excel support not available (openpyxl missing)"
        ) from exc

    wb = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
        if len(rows) > max_rows + 1:  # header + max_rows data rows
            raise HTTPException(
                status_code=413,
                detail=f"Excel sheet exceeds maximum of {max_rows} data rows",
            )
    if len(rows) < 2:
        return []

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {}
        for i, header in enumerate(headers):
            if header:
                record[header] = row[i] if i < len(row) else None
        result.append(record)
    return result


def _build_lens(row: dict[str, Any], session: Session) -> LensCatalog:
    return LensCatalog(
        manufacturer_id=_validated_manufacturer_id(session, row.get("manufacturer_id")),
        model=_str(row.get("model")),
        category=(_str(row.get("category")) or "unknown"),
        status=_str(row.get("status")) or "active",
        focal_length_mm=_float(row.get("focal_length_mm")),
        focal_length_min=_float(row.get("focal_length_min")),
        focal_length_max=_float(row.get("focal_length_max")),
        max_aperture=_float(row.get("max_aperture")),
        min_aperture=_float(row.get("min_aperture")),
        image_circle_mm=_float(row.get("image_circle_mm")),
        min_working_distance_mm=_float(row.get("min_working_distance_mm")),
        max_working_distance_mm=_float(row.get("max_working_distance_mm")),
        nominal_wd_mm=_float(row.get("nominal_wd_mm")),
        mount_type=_str(row.get("mount_type")),
        length_mm=_float(row.get("length_mm")),
        weight_g=_float(row.get("weight_g")),
        price_usd=_float(row.get("price_usd")),
        na=_float(row.get("na")),
        wavelength_min_nm=_int(row.get("wavelength_min_nm")),
        wavelength_max_nm=_int(row.get("wavelength_max_nm")),
        distortion_percent=_float(row.get("distortion_percent")),
        mtf50_lpmm=_float(row.get("mtf50_lpmm")),
        data_source="user",
        data_quality_score=1.0,
        verified=False,
    )


def _build_detector(row: dict[str, Any], session: Session) -> DetectorCatalog:
    return DetectorCatalog(
        manufacturer_id=_validated_manufacturer_id(session, row.get("manufacturer_id")),
        model=_str(row.get("model")),
        category=(_str(row.get("category")) or "unknown"),
        sensor_format_inch=_str(row.get("sensor_format_inch")),
        sensor_w_mm=_float(row.get("sensor_w_mm")),
        sensor_h_mm=_float(row.get("sensor_h_mm")),
        sensor_diag_mm=_float(row.get("sensor_diag_mm")),
        resolution_w=_int(row.get("resolution_w")),
        resolution_h=_int(row.get("resolution_h")),
        pixel_size_um=_float(row.get("pixel_size_um")),
        mount_type=_str(row.get("mount_type")),
        data_interface=_str(row.get("data_interface")),
        max_fps_full=_float(row.get("max_fps_full")),
        price_usd=_float(row.get("price_usd")),
        netd_mk=_float(row.get("netd_mk")),
        spectral_range_min_um=_float(row.get("spectral_range_min_um")),
        spectral_range_max_um=_float(row.get("spectral_range_max_um")),
        data_source="user",
        data_quality_score=1.0,
        verified=False,
    )


def import_from_upload(
    filename: str,
    content: bytes,
    session: Session,
    kind: str | None = None,
) -> dict[str, Any]:
    """Import a CSV or Excel file as user catalog data.

    Args:
        filename: Original uploaded filename.
        content: Raw file bytes.
        session: Database session.
        kind: "lenses" or "detectors". If None, inferred from filename.

    Returns:
        Summary dict with inserted counts and errors.
    """
    filename_lower = (filename or "").lower()
    if kind is None:
        if "lens" in filename_lower:
            kind = "lenses"
        elif "detector" in filename_lower or "camera" in filename_lower:
            kind = "detectors"
        else:
            raise HTTPException(
                status_code=400,
                detail="Cannot infer catalog type from filename; use lenses*.csv or detectors*.csv",
            )

    if kind not in ("lenses", "detectors"):
        raise HTTPException(status_code=400, detail="kind must be 'lenses' or 'detectors'")

    if filename_lower.endswith(".csv"):
        rows = _rows_from_csv(content)
    elif filename_lower.endswith(".xlsx"):
        rows = _rows_from_excel(content)
    else:
        raise HTTPException(
            status_code=400, detail="Only CSV and Excel (.xlsx) files are supported"
        )

    if not rows:
        return {"kind": kind, "inserted": 0, "skipped": 0, "errors": ["No data rows found"]}

    inserted = 0
    skipped = 0
    errors: list[str] = []

    if kind == "lenses":
        existing = {
            (m_id, model)
            for m_id, model in session.query(
                LensCatalog.manufacturer_id, LensCatalog.model
            ).all()
        }
        for idx, row in enumerate(rows, start=2):
            model = _str(row.get("model"))
            if not model:
                errors.append(f"Row {idx}: missing model")
                continue
            m_id = _int(row.get("manufacturer_id"))
            if m_id is None:
                m_id = _manufacturer_id_for_name(session, _str(row.get("manufacturer_name")))
            if (m_id, model) in existing:
                skipped += 1
                continue
            try:
                row["manufacturer_id"] = m_id
                item = _build_lens(row, session)
                session.add(item)
                existing.add((m_id, model))
                inserted += 1
            except Exception as exc:  # pragma: no cover
                errors.append(f"Row {idx}: {exc}")
    else:
        existing = {
            (m_id, model)
            for m_id, model in session.query(
                DetectorCatalog.manufacturer_id, DetectorCatalog.model
            ).all()
        }
        for idx, row in enumerate(rows, start=2):
            model = _str(row.get("model"))
            if not model:
                errors.append(f"Row {idx}: missing model")
                continue
            m_id = _int(row.get("manufacturer_id"))
            if m_id is None:
                m_id = _manufacturer_id_for_name(session, _str(row.get("manufacturer_name")))
            if (m_id, model) in existing:
                skipped += 1
                continue
            try:
                row["manufacturer_id"] = m_id
                item = _build_detector(row, session)
                session.add(item)
                existing.add((m_id, model))
                inserted += 1
            except Exception as exc:  # pragma: no cover
                errors.append(f"Row {idx}: {exc}")

    session.commit()
    return {
        "kind": kind,
        "inserted": inserted,
        "skipped": skipped,
        "errors": errors,
    }
